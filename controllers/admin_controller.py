# engentoria/controllers/admin_controller.py

# Importações de módulos de modelos de dados
from models import usuario_model, imobiliaria_model, imovel_model, agenda_model
# Importações de utilitários (validadores e helpers)
from utils import validators, helpers
# Importações para tipagem estática, melhorando a legibilidade e manutenção do código
from typing import Dict, Any, Optional, List, Union, Tuple
# Importação da biblioteca pandas para manipulação de dados, especialmente para relatórios
import pandas as pd
# Importação do módulo os para interagir com o sistema operacional (ex: criar diretórios)
import os
# Importações específicas para estilização de arquivos Excel gerados com openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Constante para o diretório onde os relatórios gerados serão salvos
REPORTS_DIR = "reports_generated"
# Verifica se o diretório de relatórios não existe e, em caso afirmativo, cria-o
if not os.path.exists(REPORTS_DIR):
    os.makedirs(REPORTS_DIR)

class AdminController:
    """
    Controlador para gerenciar funcionalidades administrativas do sistema.

    Esta classe encapsula a lógica de negócios para operações como:
    - Gerenciamento de usuários vistoriadores (cadastro, listagem, remoção).
    - Gerenciamento de clientes (cadastro, listagem).
    - Gerenciamento de imobiliárias (cadastro, listagem, remoção).
    - Gerenciamento de horários fixos e avulsos de vistoriadores.
    - Marcação de vistorias como improdutivas.
    - Geração de relatórios em formato Excel.
    """

    def __init__(self):
        """
        Construtor da classe AdminController.
        Neste momento, não requer inicializações complexas.
        """
        pass # Nenhuma inicialização específica necessária no momento

    # --- Seção: Gerenciamento de Vistoriadores (Usuários do tipo 'vistoriador') ---
    def cadastrar_novo_vistoriador(self, nome: str, email: str, senha: str, confirma_senha: str,
                                   telefone1: Optional[str] = None, telefone2: Optional[str] = None) -> Dict[str, Any]:
        """
        Cadastra um novo usuário do tipo 'vistoriador' no sistema.

        Realiza validações dos dados de entrada antes de prosseguir com o cadastro.
        Os campos nome, email, senha e confirmação de senha são obrigatórios.
        O email deve ter um formato válido.
        A senha deve coincidir com a confirmação e atender a critérios mínimos de segurança.
        Os telefones, se fornecidos, também são validados.

        Args:
            nome (str): Nome completo do vistoriador.
            email (str): Endereço de e-mail do vistoriador (usado para login).
            senha (str): Senha de acesso.
            confirma_senha (str): Confirmação da senha.
            telefone1 (Optional[str]): Telefone principal do vistoriador.
            telefone2 (Optional[str]): Telefone secundário do vistoriador.

        Returns:
            Dict[str, Any]: Um dicionário contendo:
                'success' (bool): True se o cadastro for bem-sucedido, False caso contrário.
                'message' (str): Mensagem informativa sobre o resultado da operação.
                'id' (Optional[int]): ID do vistoriador cadastrado, se bem-sucedido.
        """
        # Validação de campos obrigatórios
        if not validators.is_not_empty(nome) or \
           not validators.is_not_empty(email) or \
           not validators.is_not_empty(senha) or \
           not validators.is_not_empty(confirma_senha):
            return {'success': False, 'message': "Nome, e-mail, senha e confirmação de senha são obrigatórios."}
        # Validação do formato do e-mail
        if not validators.is_valid_email(email):
            return {'success': False, 'message': "Formato de e-mail inválido."}
        # Validação de confirmação de senha
        if senha != confirma_senha:
            return {'success': False, 'message': "As senhas não coincidem."}
        # Validação de complexidade/tamanho da senha
        if not validators.is_valid_password(senha, min_length=6):
            return {'success': False, 'message': "A senha deve ter pelo menos 6 caracteres."}
        # Validação do formato do telefone1, se fornecido
        if telefone1 and not validators.is_valid_phone(telefone1, allow_empty=False):
             return {'success': False, 'message': "Formato de Telefone Principal inválido."}
        # Validação do formato do telefone2, se fornecido
        if telefone2 and not validators.is_valid_phone(telefone2, allow_empty=False):
             return {'success': False, 'message': "Formato de Telefone Secundário inválido."}

        # Tentativa de cadastrar o usuário através do modelo
        vistoriador_id = usuario_model.cadastrar_usuario(
            nome=nome, email=email, senha=senha, tipo='vistoriador', # 'tipo' define o usuário como vistoriador
            telefone1=telefone1, telefone2=telefone2
        )

        # Retorno baseado no sucesso ou falha do cadastro no modelo
        if vistoriador_id:
            return {'success': True, 'message': f"Vistoriador '{nome}' cadastrado com sucesso! ID: {vistoriador_id}", 'id': vistoriador_id}
        else:
            # Mensagem genérica; o modelo pode logar detalhes específicos do erro (ex: e-mail já existe)
            return {'success': False, 'message': "Não foi possível cadastrar o vistoriador. Verifique os dados ou o console para mais detalhes."}

    def listar_todos_vistoriadores(self) -> List[Dict[str, Any]]:
        """
        Lista todos os usuários cadastrados no sistema que são do tipo 'vistoriador'.

        Returns:
            List[Dict[str, Any]]: Uma lista de dicionários, onde cada dicionário
                                  representa um vistoriador e contém seus dados.
                                  Retorna uma lista vazia se nenhum vistoriador for encontrado.
        """
        # Busca todos os usuários do tipo 'vistoriador' através do modelo
        return usuario_model.listar_usuarios_por_tipo('vistoriador')

    def remover_vistoriador(self, vistoriador_id: int) -> Dict[str, Any]:
        """
        Remove um vistoriador do sistema com base no seu ID.

        Args:
            vistoriador_id (int): O ID do vistoriador a ser removido.

        Returns:
            Dict[str, Any]: Um dicionário contendo:
                'success' (bool): True se a remoção for bem-sucedida, False caso contrário.
                'message' (str): Mensagem informativa sobre o resultado da operação.
        """
        # Validação do tipo do ID do vistoriador
        if not isinstance(vistoriador_id, int):
            return {'success': False, 'message': "ID do vistoriador inválido."}

        # Tentativa de deletar o usuário através do modelo
        sucesso = usuario_model.deletar_usuario(vistoriador_id)

        # Retorno baseado no sucesso ou falha da remoção no modelo
        if sucesso:
            return {'success': True, 'message': f"Vistoriador ID {vistoriador_id} removido com sucesso."}
        else:
            # A falha pode ocorrer se o ID não existir ou se houver restrições (ex: vistorias associadas)
            return {'success': False, 'message': f"Não foi possível remover o vistoriador ID {vistoriador_id}. Verifique se o ID é válido ou se há dependências."}

    # --- Seção: Gerenciamento de Clientes (Usuários do tipo 'cliente') ---
    def cadastrar_novo_cliente(self, nome: str, email: str,
                               telefone1: Optional[str] = None, telefone2: Optional[str] = None,
                               saldo_devedor: float = 0.0) -> Dict[str, Any]:
        """
        Cadastra um novo cliente no sistema.

        Clientes são entidades para as quais as vistorias são realizadas.
        Valida os campos obrigatórios (nome, email) e formatos (email, telefones, saldo devedor).

        Args:
            nome (str): Nome completo do cliente.
            email (str): Endereço de e-mail do cliente.
            telefone1 (Optional[str]): Telefone principal do cliente.
            telefone2 (Optional[str]): Telefone secundário do cliente.
            saldo_devedor (float): Saldo devedor inicial do cliente (padrão 0.0).

        Returns:
            Dict[str, Any]: Um dicionário contendo:
                'success' (bool): True se o cadastro for bem-sucedido, False caso contrário.
                'message' (str): Mensagem informativa sobre o resultado da operação.
                'id' (Optional[int]): ID do cliente cadastrado, se bem-sucedido.
        """
        # Validação de campos obrigatórios
        if not validators.is_not_empty(nome) or not validators.is_not_empty(email):
            return {'success': False, 'message': "Nome e e-mail do cliente são obrigatórios."}
        # Validação do formato do e-mail
        if not validators.is_valid_email(email):
            return {'success': False, 'message': "Formato de e-mail inválido."}
        # Validação do formato do telefone1, se fornecido
        if telefone1 and not validators.is_valid_phone(telefone1, allow_empty=False):
             return {'success': False, 'message': "Formato de Telefone Principal inválido."}
        # Validação do formato do telefone2, se fornecido
        if telefone2 and not validators.is_valid_phone(telefone2, allow_empty=False):
             return {'success': False, 'message': "Formato de Telefone Secundário inválido."}
        # Validação do saldo devedor
        if not isinstance(saldo_devedor, (int, float)) or saldo_devedor < 0:
            return {'success': False, 'message': "Saldo devedor deve ser um número não negativo."}

        # Tentativa de cadastrar o cliente através do modelo
        cliente_id = usuario_model.cadastrar_cliente(
            nome=nome, email=email, telefone1=telefone1, telefone2=telefone2, saldo_devedor_total=saldo_devedor
        )

        # Retorno baseado no sucesso ou falha do cadastro no modelo
        if cliente_id:
            return {'success': True, 'message': f"Cliente '{nome}' cadastrado com sucesso! ID: {cliente_id}", 'id': cliente_id}
        else:
            return {'success': False, 'message': "Não foi possível cadastrar o cliente."}

    def listar_todos_clientes_admin(self) -> List[Dict[str, Any]]:
        """
        Lista todos os clientes cadastrados no sistema.

        Returns:
            List[Dict[str, Any]]: Uma lista de dicionários, onde cada dicionário
                                  representa um cliente e contém seus dados.
                                  Retorna uma lista vazia se nenhum cliente for encontrado.
        """
        # Busca todos os clientes através do modelo
        return usuario_model.listar_todos_clientes()

    # --- Seção: Gerenciamento de Imobiliárias ---
    def cadastrar_nova_imobiliaria(self, nome: str, valor_sem_mobilia: str,
                                   valor_semi_mobiliado: str, valor_mobiliado: str) -> Dict[str, Any]:
        """
        Cadastra uma nova imobiliária no sistema.

        Imobiliárias podem ter valores por m² diferenciados para cálculo de custos de vistoria.
        Valida o nome (obrigatório) e os formatos dos valores (devem ser numéricos não negativos).

        Args:
            nome (str): Nome da imobiliária.
            valor_sem_mobilia (str): Valor por m² para imóveis sem mobília (formato string, ex: "10.50").
            valor_semi_mobiliado (str): Valor por m² para imóveis semi-mobiliados (formato string).
            valor_mobiliado (str): Valor por m² para imóveis mobiliados (formato string).

        Returns:
            Dict[str, Any]: Um dicionário contendo:
                'success' (bool): True se o cadastro for bem-sucedido, False caso contrário.
                'message' (str): Mensagem informativa sobre o resultado da operação.
                'id' (Optional[int]): ID da imobiliária cadastrada, se bem-sucedido.
        """
        # Validação do nome da imobiliária
        if not validators.is_not_empty(nome):
            return {'success': False, 'message': "Nome da imobiliária é obrigatório."}
        
        # Variáveis para armazenar os valores convertidos
        val_sm, val_smm, val_m = None, None, None
        try:
            # Preparação e conversão dos valores de string para float
            # Substitui vírgula por ponto para consistência decimal e remove espaços extras
            val_sm_str = valor_sem_mobilia.replace(',', '.').strip()
            val_smm_str = valor_semi_mobiliado.replace(',', '.').strip()
            val_m_str = valor_mobiliado.replace(',', '.').strip()

            # Validação se os valores são números positivos (ou zero)
            if not validators.is_positive_float_or_int(val_sm_str, allow_zero=True) or \
               not validators.is_positive_float_or_int(val_smm_str, allow_zero=True) or \
               not validators.is_positive_float_or_int(val_m_str, allow_zero=True):
                return {'success': False, 'message': "Valores por m² devem ser números não negativos. Use '.' como decimal."}
            
            # Conversão final para float
            val_sm = float(val_sm_str)
            val_smm = float(val_smm_str)
            val_m = float(val_m_str)
        except ValueError:
            # Captura erro se a conversão para float falhar (ex: texto não numérico)
             return {'success': False, 'message': "Formato numérico inválido para os valores por m²."}

        # Tentativa de cadastrar a imobiliária através do modelo
        imobiliaria_id = imobiliaria_model.cadastrar_imobiliaria(nome, val_sm, val_smm, val_m)
        
        # Retorno baseado no sucesso ou falha da operação no modelo
        if imobiliaria_id:
            return {'success': True, 'message': f"Imobiliária '{nome}' cadastrada com sucesso! ID: {imobiliaria_id}", 'id': imobiliaria_id}
        else:
            # A falha pode ocorrer se o nome da imobiliária já existir (assumindo que é único)
            return {'success': False, 'message': "Não foi possível cadastrar a imobiliária. O nome pode já existir."}

    def listar_todas_imobiliarias_admin(self) -> List[Dict[str, Any]]:
        """
        Lista todas as imobiliárias cadastradas no sistema.

        Returns:
            List[Dict[str, Any]]: Uma lista de dicionários, onde cada dicionário
                                  representa uma imobiliária e contém seus dados.
                                  Retorna uma lista vazia se nenhuma imobiliária for encontrada.
        """
        # Busca todas as imobiliárias através do modelo
        return imobiliaria_model.listar_todas_imobiliarias()

    def remover_imobiliaria(self, imobiliaria_id: int) -> Dict[str, Any]:
        """
        Remove uma imobiliária do sistema com base no seu ID.

        Args:
            imobiliaria_id (int): O ID da imobiliária a ser removida.

        Returns:
            Dict[str, Any]: Um dicionário contendo:
                'success' (bool): True se a remoção for bem-sucedida, False caso contrário.
                'message' (str): Mensagem informativa sobre o resultado da operação.
        """
        # Validação do tipo do ID da imobiliária
        if not isinstance(imobiliaria_id, int):
            return {'success': False, 'message': "ID da imobiliária inválido."}

        # Tentativa de deletar a imobiliária através do modelo
        sucesso = imobiliaria_model.deletar_imobiliaria(imobiliaria_id)

        # Retorno baseado no sucesso ou falha da remoção
        if sucesso:
            return {'success': True, 'message': f"Imobiliária ID {imobiliaria_id} removida com sucesso."}
        else:
            # A falha pode ocorrer se o ID não existir ou se houver imóveis associados a esta imobiliária
            return {'success': False, 'message': f"Não foi possível remover a imobiliária ID {imobiliaria_id}. Verifique se há imóveis associados."}

    # --- Seção: Gerenciamento de Horários Fixos de Vistoriadores ---
    def adicionar_horarios_fixos_para_vistoriador(self, vistoriador_id: int, dias_semana: List[str], horarios_str_lista: List[str]) -> Dict[str, Any]:
        """
        Adiciona ou atualiza os horários de trabalho fixos para um vistoriador.

        Estes horários são usados para gerar a agenda base de disponibilidade do vistoriador.
        Valida o ID do vistoriador, a lista de dias da semana e a lista de horários.

        Args:
            vistoriador_id (int): ID do vistoriador.
            dias_semana (List[str]): Lista de nomes dos dias da semana (ex: ["Segunda-feira", "Terça-feira"]).
                                     Serão convertidos para números (0-6) no modelo.
            horarios_str_lista (List[str]): Lista de horários no formato "HH:MM" (ex: ["09:00", "14:30"]).

        Returns:
            Dict[str, Any]: Dicionário com o resultado da operação.
                'success' (bool): True se bem-sucedido.
                'message' (str): Mensagem de status.
        """
        # Validações de entrada
        if not vistoriador_id or not isinstance(vistoriador_id, int) or vistoriador_id <= 0:
            return {'success': False, 'message': "ID do vistoriador inválido."}
        if not dias_semana or not isinstance(dias_semana, list) or not all(isinstance(d, str) for d in dias_semana):
            return {'success': False, 'message': "Lista de dias da semana inválida."}
        if not horarios_str_lista or not isinstance(horarios_str_lista, list) or not all(isinstance(h, str) for h in horarios_str_lista):
            return {'success': False, 'message': "Lista de horários inválida."}

        # Validação do formato de cada horário na lista
        horarios_validos = []
        for h_str in horarios_str_lista:
            if not validators.is_valid_date_format(h_str, "%H:%M", allow_empty=False):
                return {'success': False, 'message': f"Formato de horário inválido: '{h_str}'. Use HH:MM."}
            horarios_validos.append(h_str)
        
        if not horarios_validos: # Se nenhum horário passou na validação
             return {'success': False, 'message': "Nenhum horário válido fornecido."}

        # Tenta cadastrar os horários fixos através do modelo
        sucesso_cadastro = agenda_model.cadastrar_horarios_fixos_vistoriador(vistoriador_id, dias_semana, horarios_validos)
        
        if sucesso_cadastro:
            # Após adicionar/atualizar horários fixos, a agenda base precisa ser regenerada
            # para refletir essas mudanças para datas futuras.
            agenda_model.gerar_agenda_baseada_em_horarios_fixos()
            return {'success': True, 'message': "Horários fixos adicionados/atualizados e agenda regenerada."}
        else:
            return {'success': False, 'message': "Nenhum novo horário fixo foi adicionado (podem já existir ou ocorreu um erro)."}

    def remover_horario_fixo_vistoriador(self, vistoriador_id: int, dia_semana_num_str: str, horario_str: str) -> Dict[str, Any]:
        """
        Remove um horário de trabalho fixo específico de um vistoriador.

        Args:
            vistoriador_id (int): ID do vistoriador.
            dia_semana_num_str (str): Número do dia da semana como string ('0' para Domingo, ..., '6' para Sábado).
                                     Esta é a representação como armazenada no banco.
            horario_str (str): Horário a ser removido, no formato "HH:MM".

        Returns:
            Dict[str, Any]: Dicionário com o resultado da operação.
        """
        # Validações de entrada
        if not vistoriador_id or not isinstance(vistoriador_id, int) or vistoriador_id <= 0:
            return {'success': False, 'message': "ID do vistoriador inválido."}
        # Validação do dia da semana (deve ser uma string numérica de 0 a 6)
        if not dia_semana_num_str or dia_semana_num_str not in ['0','1','2','3','4','5','6']:
             return {'success': False, 'message': "Dia da semana inválido."}
        # Validação do formato do horário
        if not validators.is_valid_date_format(horario_str, "%H:%M", allow_empty=False):
            return {'success': False, 'message': f"Formato de horário inválido: '{horario_str}'. Use HH:MM."}

        # Tenta remover o horário fixo através do modelo
        sucesso = agenda_model.remover_horario_fixo_especifico(vistoriador_id, dia_semana_num_str, horario_str)
        if sucesso:
            # Idealmente, após remover um horário fixo, a agenda futura que dependia dele
            # deveria ser ajustada. A implementação atual pode não fazer isso automaticamente,
            # ou pode depender da lógica em `gerar_agenda_baseada_em_horarios_fixos` ser chamada.
            return {'success': True, 'message': "Horário fixo removido com sucesso."}
        else:
            return {'success': False, 'message': "Não foi possível remover o horário fixo (pode não existir)."}

    def listar_horarios_fixos_de_vistoriador(self, vistoriador_id: int) -> List[Dict[str, str]]:
        """
        Lista todos os horários de trabalho fixos cadastrados para um vistoriador específico.

        Args:
            vistoriador_id (int): ID do vistoriador.

        Returns:
            List[Dict[str, str]]: Lista de dicionários, cada um representando um horário fixo
                                  com 'dia_semana' (número) e 'horario' (HH:MM).
                                  Retorna lista vazia se o ID for inválido ou não houver horários.
        """
        # Validação do ID do vistoriador
        if not vistoriador_id or not isinstance(vistoriador_id, int) or vistoriador_id <= 0:
            # Log interno, não retorna erro para a view necessariamente, apenas lista vazia.
            print("❌ ID do vistoriador inválido para listar horários fixos.")
            return []
        # Busca os horários fixos através do modelo
        return agenda_model.listar_horarios_fixos_por_vistoriador(vistoriador_id)

    # --- Seção: Gerenciamento de Horários Avulsos de Vistoriadores ---
    def adicionar_horario_avulso_para_vistoriador(self, vistoriador_id: int, data_str_ddmmyyyy: str, hora_str: str) -> Dict[str, Any]:
        """
        Adiciona uma entrada de agenda avulsa (disponibilidade única) para um vistoriador.

        Permite que um vistoriador adicione um horário de trabalho específico em uma data
        que não necessariamente faz parte de seus horários fixos.

        Args:
            vistoriador_id (int): ID do vistoriador.
            data_str_ddmmyyyy (str): Data do horário avulso no formato "DD/MM/AAAA".
            hora_str (str): Hora do horário avulso no formato "HH:MM".

        Returns:
            Dict[str, Any]: Dicionário com o resultado da operação.
        """
        # Validação do ID do vistoriador
        if not vistoriador_id or not isinstance(vistoriador_id, int) or vistoriador_id <= 0:
            return {'success': False, 'message': "ID do vistoriador inválido."}
        
        # Conversão da data do formato DD/MM/AAAA para YYYY-MM-DD (formato do banco)
        data_db_format = helpers.formatar_data_para_banco(data_str_ddmmyyyy)
        if not data_db_format: # Se a conversão falhar
            return {'success': False, 'message': "Formato de data inválido. Use DD/MM/AAAA."}
        
        # Validação do formato da hora
        if not validators.is_valid_date_format(hora_str, "%H:%M", allow_empty=False):
            return {'success': False, 'message': f"Formato de hora inválido: '{hora_str}'. Use HH:MM."}

        # Tenta adicionar a entrada de agenda única através do modelo
        sucesso, mensagem = agenda_model.adicionar_entrada_agenda_unica(vistoriador_id, data_db_format, hora_str)
        return {'success': sucesso, 'message': mensagem}

    # --- Seção: Vistorias Improdutivas ---
    def marcar_vistoria_como_improdutiva(self,
                                         agenda_id: int,
                                         cliente_id: int,
                                         imovel_id: Optional[int],
                                         imobiliaria_id: Optional[int],
                                         data_vistoria_original: str, 
                                         horario_vistoria_original: str,
                                         motivo: str,
                                         valor_cobranca: float) -> Dict[str, Any]:
        """
        Registra uma vistoria que foi agendada mas não pôde ser realizada (improdutiva).

        Isso geralmente ocorre por motivos atribuíveis ao cliente (ex: não comparecimento).
        A função registra a improdutividade e pode associar uma cobrança ao cliente.

        Args:
            agenda_id (int): ID da entrada original na agenda que se tornou improdutiva.
            cliente_id (int): ID do cliente responsável pela vistoria.
            imovel_id (Optional[int]): ID do imóvel associado (se aplicável).
            imobiliaria_id (Optional[int]): ID da imobiliária associada (se aplicável).
            data_vistoria_original (str): Data original da vistoria (formato "YYYY-MM-DD").
            horario_vistoria_original (str): Horário original da vistoria (formato "HH:MM").
            motivo (str): Descrição do motivo da improdutividade.
            valor_cobranca (float): Valor a ser cobrado do cliente pela improdutividade.

        Returns:
            Dict[str, Any]: Dicionário com o resultado da operação.
        """
        # Validação de campos obrigatórios
        if not all([agenda_id, cliente_id, motivo]): # Verifica se todos são "truthy"
            return {'success': False, 'message': "ID da agenda, ID do cliente e motivo são obrigatórios."}
        # Validação do valor da cobrança
        if not isinstance(valor_cobranca, (int, float)) or valor_cobranca < 0:
            return {'success': False, 'message': "Valor da cobrança deve ser um número não negativo."}
        # Validação do formato da data original da vistoria
        if not validators.is_valid_date_format(data_vistoria_original, "%Y-%m-%d", allow_empty=False):
            return {'success': False, 'message': "Formato da data da vistoria original inválido. Esperado YYYY-MM-DD."}
        # Validação do formato do horário original da vistoria
        if not validators.is_valid_date_format(horario_vistoria_original, "%H:%M", allow_empty=False):
            return {'success': False, 'message': "Formato do horário da vistoria original inválido. Esperado HH:MM."}

        # Tenta registrar a vistoria improdutiva através do modelo
        sucesso, mensagem = agenda_model.registrar_vistoria_improdutiva(
            agenda_id_original=agenda_id,
            cliente_id=cliente_id,
            imovel_id=imovel_id,
            imobiliaria_id=imobiliaria_id,
            data_vistoria_original_str=data_vistoria_original,
            horario_vistoria_original_str=horario_vistoria_original,
            motivo=motivo,
            valor_cobranca=valor_cobranca
        )
        return {'success': sucesso, 'message': mensagem}

    # --- Seção: Geração de Relatórios ---
    def _gerar_e_formatar_relatorio_excel(self, df: pd.DataFrame, nome_base_arquivo: str, sheet_name: str = "Relatorio") -> Tuple[bool, str]:
        """
        Método auxiliar privado para gerar um arquivo Excel a partir de um DataFrame pandas e aplicar formatação.

        Este método lida com a criação do arquivo, escrita dos dados e estilização
        do cabeçalho, células, bordas e formatação de números/datas.

        Args:
            df (pd.DataFrame): DataFrame contendo os dados para o relatório.
            nome_base_arquivo (str): Nome base para o arquivo Excel (sem a extensão .xlsx).
            sheet_name (str): Nome da planilha dentro do arquivo Excel.

        Returns:
            Tuple[bool, str]: Uma tupla contendo:
                - bool: True se o relatório foi gerado com sucesso, False caso contrário.
                - str: Mensagem indicando o resultado ou o caminho do arquivo gerado.
        """
        # Verifica se o DataFrame está vazio; se sim, não há o que gerar.
        if df.empty:
            return False, "Nenhum dado encontrado para gerar o relatório."
        
        # Garante que o diretório de relatórios exista
        if not os.path.exists(REPORTS_DIR):
            try:
                os.makedirs(REPORTS_DIR)
            except OSError as e:
                return False, f"Erro ao criar diretório de relatórios: {e}"

        # Monta o caminho completo para o arquivo Excel
        caminho_arquivo = os.path.join(REPORTS_DIR, f"{nome_base_arquivo}.xlsx")
        
        try:
            # Utiliza o ExcelWriter do pandas com o engine 'openpyxl' para escrever o DataFrame
            with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False) # index=False para não escrever o índice do DataFrame no Excel
                worksheet = writer.sheets[sheet_name] # Obtém a planilha para aplicar formatação

                # --- Definição de Estilos ---
                # Estilo para o cabeçalho das colunas
                header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Cor de preenchimento azul
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) # Centralizado e com quebra de texto
                
                # Estilo para as bordas das células
                thin_border_side = Side(border_style="thin", color="000000") # Borda fina preta
                cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
                
                # Estilos de alinhamento para os dados
                data_alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
                data_alignment_right = Alignment(horizontal="right", vertical="center", wrap_text=True) # Para números
                data_alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True) # Para datas

                # Estilo para preenchimento de linhas pares (efeito zebrado)
                even_row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Cor de preenchimento azul claro

                # --- Aplicação dos Estilos ---
                # Ajusta a altura da linha do cabeçalho
                worksheet.row_dimensions[1].height = 25 
                # Itera sobre as células do cabeçalho (primeira linha) para aplicar estilos
                for col_idx, column_cell in enumerate(worksheet[1]):
                    column_cell.font = header_font
                    column_cell.fill = header_fill
                    column_cell.alignment = header_alignment
                    column_cell.border = cell_border

                # Itera sobre as linhas de dados (a partir da segunda linha)
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column), start=2):
                    worksheet.row_dimensions[row_idx].height = 20 # Ajusta a altura das linhas de dados
                    current_fill = even_row_fill if (row_idx -1) % 2 == 0 else None # Define preenchimento para linhas pares

                    # Itera sobre as células de cada linha de dados
                    for cell in row:
                        cell.border = cell_border # Aplica borda
                        if current_fill:
                            cell.fill = current_fill # Aplica preenchimento se for linha par
                        
                        # Obtém o nome da coluna para aplicar formatação condicional
                        column_name = worksheet.cell(row=1, column=cell.column).value
                        
                        # Trata valores nulos (None ou NaN) em colunas numéricas, convertendo para 0.0
                        if pd.isna(cell.value) and ("Valor" in str(column_name) or "R$" in str(column_name)):
                            cell.value = 0.0 
                        
                        # Aplica formatação e alinhamento com base no tipo de dado e nome da coluna
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = data_alignment_right # Alinha números à direita
                            if "Valor" in str(column_name) or "R$" in str(column_name) :
                                cell.number_format = 'R$ #,##0.00' # Formato monetário
                            elif isinstance(cell.value, float): 
                                cell.number_format = '#,##0.00' # Formato numérico com duas casas decimais
                        elif "Data" in str(column_name):
                            cell.alignment = data_alignment_center # Centraliza datas
                            # A formatação de data do openpyxl pode ser 'YYYY-MM-DD' ou outra conforme necessidade
                            # Se as datas já estão como objetos datetime, o Excel geralmente as formata bem.
                            # Se são strings, certificar que estão no formato correto ou converter antes.
                        else:
                            cell.alignment = data_alignment_left # Alinha outros tipos de dados à esquerda
                
                # Ajusta a largura das colunas automaticamente com base no conteúdo
                for col_idx_plus_1, column_cells in enumerate(worksheet.columns, start=1):
                    try:
                        max_len = 0
                        column_letter = get_column_letter(col_idx_plus_1) # Obtém a letra da coluna (A, B, C...)
                        
                        # Considera o tamanho do texto do cabeçalho
                        header_value = str(worksheet.cell(row=1, column=col_idx_plus_1).value)
                        max_len = len(header_value)

                        # Itera sobre as células da coluna para encontrar o maior comprimento de texto
                        for cell in column_cells:
                            if cell.value:
                                cell_text_len = len(str(cell.value))
                                # Lógica para calcular o comprimento de números formatados (monetário, decimal)
                                if cell.number_format and 'R$' in cell.number_format and isinstance(cell.value, (int,float)):
                                    formatted_value_str = f"R$ {cell.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") # Simula formatação para obter comprimento
                                    cell_text_len = len(formatted_value_str)
                                elif cell.number_format and cell.number_format == '#,##0.00' and isinstance(cell.value, float):
                                    formatted_value_str = f"{cell.value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                                    cell_text_len = len(formatted_value_str)

                                max_len = max(max_len, cell_text_len)
                        
                        # Define a largura da coluna, com um mínimo e máximo para evitar colunas muito estreitas ou largas demais
                        adjusted_width = min(max(max_len + 2, 12), 60) # Adiciona um pequeno padding (+2), mínimo de 12, máximo de 60
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    except Exception as e_col:
                        # Em caso de erro ao ajustar uma coluna, imprime um aviso e continua
                        print(f"Aviso: Erro ao ajustar coluna {get_column_letter(col_idx_plus_1)}: {e_col}")
                        pass # Continua para a próxima coluna
            
            return True, f"Relatório gerado com sucesso: {caminho_arquivo}"
        except Exception as e:
            # Captura exceções gerais durante a escrita do arquivo Excel
            return False, f"Erro ao salvar relatório em Excel: {e}"

    def gerar_relatorio_vistorias(self, tipo_relatorio_vistoria: str, data_inicio: str, data_fim: str,
                                  id_especifico: Optional[int] = None, nome_especifico: Optional[str] = None,
                                  tipo_id_especifico: Optional[str] = None) -> Dict[str, Any]:
        """
        Gera um relatório de vistorias (entrada ou saída) em formato Excel.

        O relatório pode ser geral, filtrado por vistoriador específico ou por imobiliária específica,
        dentro de um período de datas.

        Args:
            tipo_relatorio_vistoria (str): "entrada" ou "saida", define o tipo de vistoria.
            data_inicio (str): Data de início do período do relatório (formato "DD/MM/AAAA").
            data_fim (str): Data de fim do período do relatório (formato "DD/MM/AAAA").
            id_especifico (Optional[int]): ID do vistoriador ou imobiliária para filtro.
            nome_especifico (Optional[str]): Nome do vistoriador ou imobiliária (usado para nomear o arquivo).
            tipo_id_especifico (Optional[str]): "vistoriador" ou "imobiliaria", indica o tipo do id_especifico.

        Returns:
            Dict[str, Any]: Dicionário com o resultado da operação:
                'success' (bool): True se o relatório foi gerado.
                'message' (str): Mensagem de status.
                'path' (Optional[str]): Caminho do arquivo gerado, se sucesso.
        """
        # Converte as datas para o formato do banco de dados (YYYY-MM-DD)
        data_inicio_db = helpers.formatar_data_para_banco(data_inicio)
        data_fim_db = helpers.formatar_data_para_banco(data_fim)

        # Validação das datas convertidas
        if not data_inicio_db or not data_fim_db:
            return {'success': False, 'message': "Formato de data inválido. Use DD/MM/YYYY."}
        
        # DataFrame que armazenará os dados do relatório
        df_relatorio = pd.DataFrame()
        # Nome base para o arquivo Excel
        nome_arquivo_base = f"relatorio_vistorias_{tipo_relatorio_vistoria}"
        
        # Lógica para buscar os dados do relatório com base nos filtros fornecidos
        if tipo_id_especifico == 'vistoriador' and id_especifico:
            # Relatório filtrado por vistoriador
            if tipo_relatorio_vistoria == "entrada":
                df_relatorio = agenda_model.obter_dados_relatorio_entrada_por_vistoriador(data_inicio_db, data_fim_db, id_especifico)
            elif tipo_relatorio_vistoria == "saida":
                df_relatorio = agenda_model.obter_dados_relatorio_saida_por_vistoriador(data_inicio_db, data_fim_db, id_especifico)
            # Adiciona informações do vistoriador ao nome do arquivo
            nome_arquivo_base += f"_vist_{nome_especifico.replace(' ', '_')}_{id_especifico}" if nome_especifico else f"_vist_{id_especifico}"
        
        elif tipo_id_especifico == 'imobiliaria' and id_especifico:
            # Relatório filtrado por imobiliária
            if tipo_relatorio_vistoria == "entrada":
                df_relatorio = agenda_model.obter_dados_relatorio_entrada_por_imobiliaria(data_inicio_db, data_fim_db, id_especifico)
            elif tipo_relatorio_vistoria == "saida":
                df_relatorio = agenda_model.obter_dados_relatorio_saida_por_imobiliaria(data_inicio_db, data_fim_db, id_especifico)
            # Adiciona informações da imobiliária ao nome do arquivo
            nome_arquivo_base += f"_imob_{nome_especifico.replace(' ', '_')}_{id_especifico}" if nome_especifico else f"_imob_{id_especifico}"
        
        else: # Relatório geral, sem filtro por ID específico
            if tipo_relatorio_vistoria == "entrada":
                df_relatorio = agenda_model.obter_dados_relatorio_entrada_geral(data_inicio_db, data_fim_db)
            elif tipo_relatorio_vistoria == "saida":
                df_relatorio = agenda_model.obter_dados_relatorio_saida_geral(data_inicio_db, data_fim_db)
            nome_arquivo_base += "_geral"
        
        # Adiciona o período de datas ao nome do arquivo
        nome_arquivo_base += f"_{data_inicio_db}_a_{data_fim_db}"

        # Chama o método auxiliar para gerar e formatar o arquivo Excel
        sucesso_excel, msg_excel = self._gerar_e_formatar_relatorio_excel(df_relatorio, nome_arquivo_base, f"Vistorias {tipo_relatorio_vistoria.title()}")
        
        if sucesso_excel:
            return {'success': True, 'message': msg_excel, 'path': os.path.join(REPORTS_DIR, f"{nome_arquivo_base}.xlsx")}
        else:
            return {'success': False, 'message': msg_excel}

    def gerar_relatorio_clientes_devedores(self,
                                           data_inicio_cancelamento: Optional[str] = None, 
                                           data_fim_cancelamento: Optional[str] = None,
                                           imobiliaria_id_filtro: Optional[int] = None,
                                           apenas_nao_pagos: bool = True) -> Dict[str, Any]:
        """
        Gera um relatório de vistorias improdutivas (que podem gerar débitos para clientes).

        O relatório pode ser filtrado por período de marcação da improdutividade,
        por imobiliária associada à vistoria original, e se deve incluir apenas
        débitos não pagos ou todos.

        Args:
            data_inicio_cancelamento (Optional[str]): Data de início do período de marcação
                                                      da improdutiva (formato "DD/MM/AAAA").
            data_fim_cancelamento (Optional[str]): Data de fim do período de marcação
                                                   da improdutiva (formato "DD/MM/AAAA").
            imobiliaria_id_filtro (Optional[int]): ID da imobiliária para filtrar as vistorias.
            apenas_nao_pagos (bool): Se True (padrão), lista apenas débitos de improdutivas
                                     que ainda não foram pagos. Se False, lista todos.

        Returns:
            Dict[str, Any]: Dicionário com o resultado da operação.
        """
        # Converte as datas de filtro para o formato do banco, se fornecidas
        data_inicio_db = helpers.formatar_data_para_banco(data_inicio_cancelamento) if data_inicio_cancelamento else None
        data_fim_db = helpers.formatar_data_para_banco(data_fim_cancelamento) if data_fim_cancelamento else None
        
        # Obtém o nome da imobiliária para usar no nome do arquivo, se um filtro for aplicado
        nome_imobiliaria_filtro_str = "Todas"
        if imobiliaria_id_filtro and imobiliaria_id_filtro > 0:
            imob_data = imobiliaria_model.obter_imobiliaria_por_id(imobiliaria_id_filtro)
            if imob_data:
                nome_imobiliaria_filtro_str = imob_data['nome']
            else:
                # Se o ID da imobiliária fornecido não for encontrado, remove o filtro
                print(f"Aviso: Imobiliária ID {imobiliaria_id_filtro} não encontrada para filtro de devedores. Listando para todas.")
                imobiliaria_id_filtro = None 
        
        # Busca os dados de clientes devedores (vistorias improdutivas) através do modelo
        df_devedores = usuario_model.obter_dados_clientes_devedores(
            data_inicio_marcacao=data_inicio_db, 
            data_fim_marcacao=data_fim_db,       
            imobiliaria_id_filtro=imobiliaria_id_filtro,
            apenas_nao_pagos=apenas_nao_pagos
        )
        
        # Bloco de DEBUG para verificar o conteúdo do DataFrame antes de gerar o Excel
        # (Pode ser removido ou comentado em produção)
        print("DEBUG: Colunas do DataFrame de Devedores:", df_devedores.columns.tolist() if not df_devedores.empty else "DataFrame Vazio")
        if not df_devedores.empty:
            print("DEBUG: Primeiras linhas do DataFrame de Devedores:")
            print(df_devedores.head().to_string())
        else:
            print("DEBUG: DataFrame de Devedores (Vistorias Improdutivas) está vazio.")

        # Monta o nome do arquivo Excel com base nos filtros aplicados
        nome_arquivo = "relatorio_vistorias_improdutivas"
        if imobiliaria_id_filtro and imobiliaria_id_filtro > 0:
            nome_arquivo += f"_{nome_imobiliaria_filtro_str.replace(' ', '_').lower()}"
        if data_inicio_db or data_fim_db:
            str_data_inicio = f"_de_{data_inicio_db}" if data_inicio_db else ""
            str_data_fim = f"_ate_{data_fim_db}" if data_fim_db else ""
            if str_data_inicio or str_data_fim:
                 nome_arquivo += f"{str_data_inicio}{str_data_fim}"
        nome_arquivo += "_pagas" if not apenas_nao_pagos else "_nao_pagas" # Indica se são pagas ou não pagas
        
        # Gera e formata o relatório Excel
        sucesso_excel, msg_excel = self._gerar_e_formatar_relatorio_excel(df_devedores, nome_arquivo, "Vistorias Improdutivas")

        if sucesso_excel:
            return {'success': True, 'message': msg_excel, 'path': os.path.join(REPORTS_DIR, f"{nome_arquivo}.xlsx")}
        else:
            return {'success': False, 'message': msg_excel}


# Bloco de execução principal (testes rápidos do controlador)
if __name__ == '__main__':
    # Instancia o controlador
    admin_ctrl = AdminController()
    print("AdminController instanciado. Descomente e adapte os testes abaixo para verificar funcionalidaes.")

    # --- Exemplos de Testes (descomentar e adaptar conforme necessário) ---

    # Teste: Cadastrar novo vistoriador
    # resultado_cad_vist = admin_ctrl.cadastrar_novo_vistoriador(
    #     nome="Vistoriador Teste Comentado",
    #     email="vist.teste.comentado@example.com",
    #     senha="password123",
    #     confirma_senha="password123",
    #     telefone1="(62)99999-8877"
    # )
    # print(f"Cadastro Vistoriador: {resultado_cad_vist}")

    # Teste: Listar vistoriadores
    # print("\n--- Vistoriadores Cadastrados ---")
    # vistoriadores = admin_ctrl.listar_todos_vistoriadores()
    # if vistoriadores:
    #     for v in vistoriadores:
    #         print(f"ID: {v['id']}, Nome: {v['nome']}, Email: {v['email']}")
    # else:
    #     print("Nenhum vistoriador cadastrado.")

    # Teste: Cadastrar nova imobiliária
    # resultado_cad_imob = admin_ctrl.cadastrar_nova_imobiliaria(
    #     nome="Imobiliária Modelo Comentada",
    #     valor_sem_mobilia="10.0",
    #     valor_semi_mobiliado="12.5",
    #     valor_mobiliado="15.0"
    # )
    # print(f"\nCadastro Imobiliária: {resultado_cad_imob}")

    # Teste: Listar imobiliárias
    # print("\n--- Imobiliárias Cadastradas ---")
    # imobiliarias = admin_ctrl.listar_todas_imobiliarias_admin()
    # if imobiliarias:
    #     for imob in imobiliarias:
    #         print(f"ID: {imob['id']}, Nome: {imob['nome']}, Valor s/m: {imob['valor_m2_sem_mobilia']}")
    # else:
    #     print("Nenhuma imobiliária cadastrada.")

    # Teste: Adicionar horários fixos (assumindo que existe um vistoriador com ID 1)
    # id_vist_para_horario = 1 # Substituir pelo ID de um vistoriador existente
    # vist_existe = any(v['id'] == id_vist_para_horario for v in admin_ctrl.listar_todos_vistoriadores())
    # if vist_existe:
    #     resultado_add_horario = admin_ctrl.adicionar_horarios_fixos_para_vistoriador(
    #         vistoriador_id=id_vist_para_horario,
    #         dias_semana=["Segunda-feira", "Quarta-feira"], # Nomes dos dias
    #         horarios_str_lista=["09:00", "10:00", "14:00", "15:00"]
    #     )
    #     print(f"\nAdicionar Horários Fixos (Vistoriador ID {id_vist_para_horario}): {resultado_add_horario}")
    # else:
    #     print(f"\nAVISO: Vistoriador com ID {id_vist_para_horario} não encontrado. Teste de horários fixos não executado.")

    # Teste: Gerar relatório de vistorias de entrada geral para um período
    # print("\n--- Gerando Relatório de Vistorias de Entrada (Exemplo) ---")
    # resultado_rel_entrada = admin_ctrl.gerar_relatorio_vistorias(
    #     tipo_relatorio_vistoria="entrada",
    #     data_inicio="01/01/2024", # Ajustar datas conforme seus dados de teste
    #     data_fim="31/12/2024"
    # )
    # print(f"Resultado Relatório Entrada: {resultado_rel_entrada}")
    # if resultado_rel_entrada['success']:
    #     print(f"Relatório salvo em: {resultado_rel_entrada.get('path')}")

    # Teste: Gerar relatório de clientes devedores (vistorias improdutivas não pagas)
    # print("\n--- Gerando Relatório de Vistorias Improdutivas Não Pagas (Exemplo) ---")
    # resultado_rel_devedores = admin_ctrl.gerar_relatorio_clientes_devedores(
    #     apenas_nao_pagos=True
    #     # data_inicio_cancelamento="01/05/2024", # Opcional
    #     # data_fim_cancelamento="31/05/2024",   # Opcional
    #     # imobiliaria_id_filtro=1              # Opcional, ID de uma imobiliária existente
    # )
    # print(f"Resultado Relatório Devedores: {resultado_rel_devedores}")
    # if resultado_rel_devedores['success']:
    #     print(f"Relatório salvo em: {resultado_rel_devedores.get('path')}")

